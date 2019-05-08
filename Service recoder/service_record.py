from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.drawing.image import Image
import os
import time

date = time.strftime("%Y%m%d", time.localtime())    # %Y%m%d 顯示格式  

wb = load_workbook(filename = 'MACHINE SERVICE RECORD.xlsx')
# sheet_ranges = wb['6MVQ00051']
ws = wb.active

img = Image('litz.jpg')
ws.add_image(img, 'G1')

agent = str.title(input('請輸入客戶(代理商):'))
agent_ref = input('請輸入客戶(代理商)代號:')
country = str.title(input('請輸入國家:'))
model = str.upper(input('請輸入機型:'))    #將小寫轉換為大寫
serial = str.upper(input('請輸入機號:'))
contact_person = str.title(input('請輸入連絡人:'))
mfg_data = input('請輸入生產日期:')
warrant = input('保固狀態(1=保固內 2=保固外):')
address = str.title(input('客戶地址:'))
processin = input('加工類型:')
ws['C3'] = agent
ws['I3'] = agent_ref
ws['N3'] = country
ws['C5'] = model
ws['I5'] = serial
ws['N5'] = time.strftime("%Y/%m/%d", time.localtime())
ws['C7'] = contact_person
ws['I7'] = mfg_data
if warrant == '1':
    ws['N7'] = '保固內'
else:
    ws['N7'] = '保固外'
ws['C9'] = address
ws['C11'] = processin


def problem():
    problem_coor = 14
    n = 0
    while n <= 5:
        grid_1 = ('A' +str(problem_coor))
        grid_2 = ('B' +str(problem_coor))     
        problem = input('問題描述(不再輸入請打na):')
        if problem == 'na' or n > 5:
            break
        n += 1
        ws[ grid_1 ] = (str(n) + '.')
        ws[ grid_2 ] = problem
        problem_coor += 2
    

def solution():
    solution_coor = 27
    n = 0
    while n <= 5:
        grid_1 = ('A' +str(solution_coor))
        grid_2 = ('B' +str(solution_coor))     
        solution = input('解決方案(不再輸入請打na):')
        if solution == 'na' or n > 5:
            break
        n += 1
        ws[ grid_1 ] = (str(n) + '.')
        ws[ grid_2 ] = solution
        solution_coor += 2    



problem()
solution()

path = model
new_name = serial
wb.save(os.path.join(path, new_name +'(' + date + ')' + ".xlsx"))


